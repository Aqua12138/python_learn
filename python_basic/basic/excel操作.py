import openpyxl
wb=openpyxl.Workbook()
sheet=wb.active
sheet.title='zhx'
sheet['A2']='我真棒'
row=['a','b','c']
sheet.append(row)
rows=[['a','b','c'],['d','e','f']]
for i in rows:
    sheet.append(i)
wb.save('first.xlsx')

wb = openpyxl.load_workbook('first.xlsx')
sheet = wb['zhx']
sheetname = wb.sheetnames
print(sheetname)
A1_cell = sheet['A2']
print(A1_cell.value)